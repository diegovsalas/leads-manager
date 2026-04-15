"""
Generador de reportes QBR (Quarterly Business Review) en Excel.

Genera un Excel profesional con 5 hojas:
  1. Resumen Ejecutivo — KPIs + desglose mensual
  2. Detalle por Zona
  3. Cobertura Sucursales — semáforo mes a mes
  4. Detalle Citas — registros limpios con autofiltro
  5. Auditoría — registros eliminados por deduplicación
"""

from datetime import datetime
from io import BytesIO
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

from models import CSAccount as Account, CSAppointment as Appointment

# ---------------------------------------------------------------------------
# Clasificación de servicios
# ---------------------------------------------------------------------------
# Servicios recurrentes mensuales: se esperan en TODAS las sucursales
SERVICIOS_RECURRENTES = {
    'Fumigación y control de plagas',
    'Póliza Fumigación y control de plagas',
    'Aroma Advance',
    'Aroma Advance Pro',
    'Aroma Home Pro',
    'Aroma One',
    'Aroma Plus Pro (Plata)',
    'Aroma Plus Pro (Negro)',
    'Aroma Extreme Pro',
}

# Servicios por evento: no aplican a todas las sucursales
SERVICIOS_EVENTO = {
    'Levantamiento Fumigación',
    'Levantamiento Aroma',
    'Fumigaciones Especiales',
    'Incidencias',
    'Incidencias Aromatex',
    'Instalacion',
    'Otro',
}

# ---------------------------------------------------------------------------
# Paleta de colores y estilos
# ---------------------------------------------------------------------------
AZUL_OSCURO = PatternFill('solid', fgColor='1E3A5F')
AZUL_MEDIO = PatternFill('solid', fgColor='2563EB')
GRIS_CLARO = PatternFill('solid', fgColor='F3F4F6')
VERDE_LIGHT = PatternFill('solid', fgColor='D1FAE5')
ROJO_LIGHT = PatternFill('solid', fgColor='FEE2E2')
AMARILLO_LIGHT = PatternFill('solid', fgColor='FEF9C3')
MORADO_FILL = PatternFill('solid', fgColor='EDE9FE')

FONT_HEADER = Font(name='Arial', bold=True, color='FFFFFF', size=11)
FONT_TITLE = Font(name='Arial', bold=True, color='FFFFFF', size=14)
FONT_SUBTITLE = Font(name='Arial', bold=True, color='1E3A5F', size=12)
FONT_NORMAL = Font(name='Arial', size=10)
FONT_BOLD = Font(name='Arial', bold=True, size=10)
FONT_SMALL = Font(name='Arial', size=9, color='666666')
FONT_GREEN = Font(name='Arial', bold=True, size=10, color='16A34A')
FONT_RED = Font(name='Arial', bold=True, size=10, color='DC2626')
FONT_KPI_VALUE = Font(name='Arial', bold=True, size=16, color='1E3A5F')
FONT_KPI_LABEL = Font(name='Arial', size=9, color='6B7280')

BORDER_THIN = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Meses del Q1
MESES_Q1 = {1: 'Enero', 2: 'Febrero', 3: 'Marzo'}


def _apply_header_row(ws, row, max_col, fill=AZUL_OSCURO):
    """Aplica estilo de encabezado a una fila."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _apply_data_row(ws, row, max_col, alt=False):
    """Aplica estilo de fila de datos."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_NORMAL
        cell.border = BORDER_THIN
        if alt:
            cell.fill = GRIS_CLARO


def _auto_width(ws, min_width=10, max_width=40):
    """Ajusta ancho de columnas al contenido."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _pct_font(value):
    """Devuelve fuente verde si >= 90%, roja si < 90%."""
    if value is None:
        return FONT_NORMAL
    return FONT_GREEN if value >= 0.9 else FONT_RED


# ---------------------------------------------------------------------------
# Lógica de datos
# ---------------------------------------------------------------------------

def _get_appointments_data(account: Account):
    """Obtiene y estructura todas las citas de la cuenta."""
    apts = (Appointment.query
            .filter_by(account_id=account.id)
            .order_by(Appointment.fecha_inicio.desc())
            .all())

    records = []
    for a in apts:
        mes = a.fecha_inicio.month if a.fecha_inicio else None
        anio = a.fecha_inicio.year if a.fecha_inicio else None
        # Solo Q1 2026
        if mes is None or anio != 2026 or mes not in (1, 2, 3):
            continue

        es_recurrente = a.titulo_servicio in SERVICIOS_RECURRENTES
        tipo = 'recurrente' if es_recurrente else 'evento'

        records.append({
            'id': a.id,
            'propiedad': a.propiedad,
            'zona': a.zona if a.zona and a.zona != 'nan' else 'Sin zona',
            'tecnico': a.tecnico if a.tecnico and a.tecnico != 'nan' else '',
            'servicio': a.titulo_servicio,
            'tipo_servicio': tipo,
            'mes': mes,
            'estatus': a.estatus,
            'fecha_inicio': a.fecha_inicio,
            'fecha_terminacion': a.fecha_terminacion,
            'cantidad': a.cantidad,
        })

    return records


def _deduplicar(records):
    """
    Deduplicación: por cada (propiedad, mes, servicio), conservar:
    - El registro más reciente con estatus 'Terminada' si existe.
    - Si no, el más reciente sin importar estatus.
    Devuelve (limpios, descartados).
    """
    # Agrupar por clave
    grupos = defaultdict(list)
    for r in records:
        key = (r['propiedad'], r['mes'], r['servicio'])
        grupos[key].append(r)

    limpios = []
    descartados = []

    for key, grupo in grupos.items():
        if len(grupo) == 1:
            limpios.append(grupo[0])
            continue

        # Ordenar por fecha descendente
        grupo.sort(key=lambda x: x['fecha_inicio'] or datetime.min, reverse=True)

        # Buscar terminadas
        terminadas = [r for r in grupo if r['estatus'] == 'Terminada']
        if terminadas:
            elegido = terminadas[0]
        else:
            elegido = grupo[0]

        limpios.append(elegido)
        for r in grupo:
            if r is not elegido:
                descartados.append(r)

    return limpios, descartados


def _clasificar_cobertura(records_sucursal):
    """
    Clasifica la cobertura de una sucursal dados sus registros limpios.
    Retorna: COMPLETO, PARCIAL, NO REALIZADO
    """
    terminadas = sum(1 for r in records_sucursal if r['estatus'] == 'Terminada')
    total = len(records_sucursal)

    if total == 0:
        return 'SIN CITA'
    if terminadas == total:
        return 'COMPLETO'
    if terminadas > 0:
        return 'PARCIAL'
    return 'NO REALIZADO'


# ---------------------------------------------------------------------------
# Generación del Excel
# ---------------------------------------------------------------------------

def generar_qbr(account: Account, trimestre: str = 'Q1 2026') -> BytesIO:
    """
    Genera el reporte QBR en Excel y lo devuelve como BytesIO.
    """
    wb = Workbook()

    # Obtener datos
    all_records = _get_appointments_data(account)
    limpios, descartados = _deduplicar(all_records)

    # Separar recurrentes y eventos
    recurrentes = [r for r in limpios if r['tipo_servicio'] == 'recurrente']
    eventos = [r for r in limpios if r['tipo_servicio'] == 'evento']

    # Portafolio = propiedades únicas con al menos un servicio recurrente
    portafolio = set(r['propiedad'] for r in recurrentes)
    # Si no hay recurrentes, usar todas las propiedades
    if not portafolio:
        portafolio = set(r['propiedad'] for r in limpios)

    # Tipos de servicio recurrente presentes
    tipos_recurrentes = sorted(set(r['servicio'] for r in recurrentes))
    tipos_eventos = sorted(set(r['servicio'] for r in eventos))

    # =====================================================================
    # HOJA 1: RESUMEN EJECUTIVO
    # =====================================================================
    ws1 = wb.active
    ws1.title = 'Resumen Ejecutivo'
    ws1.sheet_properties.tabColor = '1E3A5F'

    # Banner
    ws1.merge_cells('A1:I1')
    ws1.cell(row=1, column=1, value=f'QBR — {account.nombre}')
    ws1['A1'].font = FONT_TITLE
    ws1['A1'].fill = AZUL_OSCURO
    ws1['A1'].alignment = ALIGN_CENTER
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells('A2:I2')
    ws1.cell(row=2, column=1, value=f'{trimestre} · {account.unidades_contratadas} · KAM: {account.kam.nombre}')
    ws1['A2'].font = Font(name='Arial', size=11, color='FFFFFF')
    ws1['A2'].fill = AZUL_MEDIO
    ws1['A2'].alignment = ALIGN_CENTER

    # KPIs principales (fila 4)
    row = 4
    kpis = [
        ('Portafolio', len(portafolio)),
        ('Sucursales\natendidas', len(set(r['propiedad'] for r in recurrentes
                                          if r['estatus'] == 'Terminada'))),
        ('Citas\ntotales', len(limpios)),
        ('Terminadas', sum(1 for r in limpios if r['estatus'] == 'Terminada')),
        ('Canceladas', sum(1 for r in limpios if r['estatus'] == 'Cancelada')),
        ('No\nRealizadas', sum(1 for r in limpios if r['estatus'] == 'No Realizada')),
    ]

    # % Cobertura y % Cumplimiento
    sucs_atendidas = len(set(r['propiedad'] for r in recurrentes
                             if r['estatus'] == 'Terminada'))
    pct_cobertura = sucs_atendidas / len(portafolio) if portafolio else 0
    terminadas_rec = sum(1 for r in recurrentes if r['estatus'] == 'Terminada')
    # Denominador = portafolio × 3 meses para recurrente mensual
    denominador_cumpl = len(portafolio) * 3
    pct_cumplimiento = terminadas_rec / denominador_cumpl if denominador_cumpl else 0

    kpis.append(('% Cobertura', pct_cobertura))
    kpis.append(('% Cumplimiento\n(vs portafolio)', pct_cumplimiento))

    for i, (label, value) in enumerate(kpis):
        col = i + 1
        cell_val = ws1.cell(row=row, column=col)
        if isinstance(value, float) and value <= 1:
            cell_val.value = value
            cell_val.number_format = '0.0%'
            cell_val.font = _pct_font(value)
        else:
            cell_val.value = value
            cell_val.font = FONT_KPI_VALUE
        cell_val.alignment = ALIGN_CENTER
        cell_val.border = BORDER_THIN

        cell_lbl = ws1.cell(row=row + 1, column=col)
        cell_lbl.value = label
        cell_lbl.font = FONT_KPI_LABEL
        cell_lbl.alignment = Alignment(horizontal='center', vertical='top',
                                        wrap_text=True)
        cell_lbl.border = BORDER_THIN

    # --- Desglose mensual por servicio recurrente ---
    row = 7
    ws1.cell(row=row, column=1, value='SERVICIOS RECURRENTES — Desglose Mensual')
    ws1['A7'].font = FONT_SUBTITLE
    ws1.merge_cells('A7:I7')

    for tipo_srv in tipos_recurrentes:
        row += 1
        ws1.cell(row=row, column=1, value=tipo_srv)
        ws1.cell(row=row, column=1).font = Font(name='Arial', bold=True,
                                                  size=10, color='1E3A5F')
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

        row += 1
        headers = ['Mes', 'Base Portafolio', 'Agendadas', 'Sin Agendar',
                    'Terminadas', 'Canceladas', 'No Realizadas',
                    '% Cumplimiento', 'Ingreso Est.']
        for c, h in enumerate(headers, 1):
            ws1.cell(row=row, column=c, value=h)
        _apply_header_row(ws1, row, len(headers), AZUL_MEDIO)

        recs_tipo = [r for r in recurrentes if r['servicio'] == tipo_srv]
        total_term = 0
        total_agendadas = 0

        for mes_num in (1, 2, 3):
            row += 1
            recs_mes = [r for r in recs_tipo if r['mes'] == mes_num]
            base = len(portafolio)
            agendadas = len(recs_mes)
            sin_agendar = max(0, base - agendadas)
            terminadas = sum(1 for r in recs_mes if r['estatus'] == 'Terminada')
            canceladas = sum(1 for r in recs_mes if r['estatus'] == 'Cancelada')
            no_realizadas = sum(1 for r in recs_mes
                                if r['estatus'] == 'No Realizada')
            pct = terminadas / base if base > 0 else 0

            total_term += terminadas
            total_agendadas += agendadas

            vals = [MESES_Q1[mes_num], base, agendadas, sin_agendar,
                    terminadas, canceladas, no_realizadas, pct, None]
            for c, v in enumerate(vals, 1):
                cell = ws1.cell(row=row, column=c, value=v)
                cell.border = BORDER_THIN
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_CENTER
                if c == 4 and v > 0:  # Sin Agendar en rojo
                    cell.font = FONT_RED
                if c == 8:  # % Cumplimiento
                    cell.number_format = '0.0%'
                    cell.font = _pct_font(v)

            _apply_data_row(ws1, row, len(headers), mes_num % 2 == 0)

        # Subtotal Q1
        row += 1
        base_q1 = len(portafolio) * 3
        pct_q1 = total_term / base_q1 if base_q1 > 0 else 0
        vals_total = ['TOTAL Q1', base_q1, total_agendadas,
                      max(0, base_q1 - total_agendadas), total_term,
                      sum(1 for r in recs_tipo if r['estatus'] == 'Cancelada'),
                      sum(1 for r in recs_tipo if r['estatus'] == 'No Realizada'),
                      pct_q1, None]
        for c, v in enumerate(vals_total, 1):
            cell = ws1.cell(row=row, column=c, value=v)
            cell.font = FONT_BOLD
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER
            if c == 8:
                cell.number_format = '0.0%'
                cell.font = _pct_font(v)

        row += 1  # Espacio

    # --- Servicios por evento ---
    if eventos:
        row += 1
        ws1.cell(row=row, column=1, value='SERVICIOS POR EVENTO')
        ws1.cell(row=row, column=1).font = FONT_SUBTITLE
        ws1.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=9)

        row += 1
        ev_headers = ['Servicio', 'Mes', 'Agendadas', 'Terminadas',
                      'Canceladas', 'No Realizadas', '% Eficacia', 'Ingreso']
        for c, h in enumerate(ev_headers, 1):
            ws1.cell(row=row, column=c, value=h)
        _apply_header_row(ws1, row, len(ev_headers),
                          PatternFill('solid', fgColor='7C3AED'))

        for tipo_srv in tipos_eventos:
            recs_ev = [r for r in eventos if r['servicio'] == tipo_srv]
            for mes_num in (1, 2, 3):
                recs_mes = [r for r in recs_ev if r['mes'] == mes_num]
                if not recs_mes:
                    continue
                row += 1
                agendadas = len(recs_mes)
                terminadas = sum(1 for r in recs_mes
                                  if r['estatus'] == 'Terminada')
                canceladas = sum(1 for r in recs_mes
                                  if r['estatus'] == 'Cancelada')
                no_real = sum(1 for r in recs_mes
                              if r['estatus'] == 'No Realizada')
                pct = terminadas / agendadas if agendadas > 0 else 0

                vals = [tipo_srv, MESES_Q1[mes_num], agendadas, terminadas,
                        canceladas, no_real, pct, None]
                for c, v in enumerate(vals, 1):
                    cell = ws1.cell(row=row, column=c, value=v)
                    cell.border = BORDER_THIN
                    cell.font = FONT_NORMAL
                    cell.alignment = ALIGN_CENTER
                    if c == 7:
                        cell.number_format = '0.0%'
                        cell.font = _pct_font(v)
                _apply_data_row(ws1, row, len(ev_headers), row % 2 == 0)

    # Notas de metodología
    row += 2
    notas = [
        f'Reporte generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        f'Registros totales procesados: {len(all_records)}',
        f'Registros tras deduplicación: {len(limpios)} '
        f'(eliminados: {len(descartados)})',
        'Deduplicación: por (Sucursal, Mes, Servicio) se conserva el '
        'registro Terminado más reciente, o el más reciente si no hay.',
        '% Cumplimiento usa el portafolio como denominador, '
        'no las citas agendadas.',
    ]
    for nota in notas:
        ws1.cell(row=row, column=1, value=nota)
        ws1.cell(row=row, column=1).font = FONT_SMALL
        ws1.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=9)
        row += 1

    _auto_width(ws1)

    # =====================================================================
    # HOJA 2: DETALLE POR ZONA
    # =====================================================================
    ws2 = wb.create_sheet('Detalle por Zona')
    ws2.sheet_properties.tabColor = '2563EB'

    # Solo recurrentes para métricas de zona
    zonas_data = defaultdict(lambda: {
        'sucursales': set(), 'atendidas': set(),
        'citas': 0, 'terminadas': 0, 'canceladas': 0, 'no_realizadas': 0
    })
    for r in recurrentes:
        z = zonas_data[r['zona']]
        z['sucursales'].add(r['propiedad'])
        z['citas'] += 1
        if r['estatus'] == 'Terminada':
            z['atendidas'].add(r['propiedad'])
            z['terminadas'] += 1
        elif r['estatus'] == 'Cancelada':
            z['canceladas'] += 1
        elif r['estatus'] == 'No Realizada':
            z['no_realizadas'] += 1

    z_headers = ['Zona', 'Sucursales', 'Atendidas', '% Cobertura',
                 'Citas', 'Terminadas', 'Cancel/NoRealiz', '% Cumplimiento']
    for c, h in enumerate(z_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _apply_header_row(ws2, 1, len(z_headers))

    row = 2
    for zona in sorted(zonas_data.keys()):
        zd = zonas_data[zona]
        n_sucs = len(zd['sucursales'])
        n_atend = len(zd['atendidas'])
        pct_cob = n_atend / n_sucs if n_sucs > 0 else 0
        pct_cumpl = zd['terminadas'] / zd['citas'] if zd['citas'] > 0 else 0

        vals = [zona, n_sucs, n_atend, pct_cob,
                zd['citas'], zd['terminadas'],
                zd['canceladas'] + zd['no_realizadas'], pct_cumpl]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = BORDER_THIN
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            if c in (4, 8):
                cell.number_format = '0.0%'
                cell.font = _pct_font(v)
        _apply_data_row(ws2, row, len(z_headers), row % 2 == 0)
        row += 1

    _auto_width(ws2)

    # =====================================================================
    # HOJA 3: COBERTURA SUCURSALES
    # =====================================================================
    ws3 = wb.create_sheet('Cobertura Sucursales')
    ws3.sheet_properties.tabColor = '16A34A'

    s_headers = ['Sucursal', 'Zona', 'Enero', 'Febrero', 'Marzo',
                 'Terminadas', 'Total Citas', 'Clasificación']
    for c, h in enumerate(s_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    _apply_header_row(ws3, 1, len(s_headers))

    # Agrupar recurrentes por sucursal
    sucs_data = defaultdict(lambda: {
        'zona': 'Sin zona',
        'meses': {1: [], 2: [], 3: []}
    })
    for r in recurrentes:
        sd = sucs_data[r['propiedad']]
        sd['zona'] = r['zona']
        sd['meses'][r['mes']].append(r['estatus'])

    # También incluir sucursales del portafolio sin citas
    for prop in portafolio:
        if prop not in sucs_data:
            sucs_data[prop] = {'zona': 'Sin zona', 'meses': {1: [], 2: [], 3: []}}

    row = 2
    for prop in sorted(sucs_data.keys()):
        sd = sucs_data[prop]
        all_recs = []
        for m in (1, 2, 3):
            all_recs.extend(sd['meses'][m])

        terminadas = sum(1 for e in all_recs if e == 'Terminada')
        total = len(all_recs)
        clasificacion = _clasificar_cobertura(
            [{'estatus': e} for e in all_recs])

        ws3.cell(row=row, column=1, value=prop).font = FONT_NORMAL
        ws3.cell(row=row, column=2, value=sd['zona']).font = FONT_NORMAL

        # Estatus por mes con semáforo (columnas 3, 4, 5 → meses 1, 2, 3)
        for mes_num in (1, 2, 3):
            col = mes_num + 2  # col 3=Enero, 4=Feb, 5=Mar
            cell = ws3.cell(row=row, column=col)
            estatuses = sd['meses'][mes_num]

            if not estatuses:
                cell.value = 'Sin cita'
                cell.fill = AMARILLO_LIGHT
            elif 'Terminada' in estatuses:
                cell.value = 'Terminada'
                cell.fill = VERDE_LIGHT
            elif any(e in ('Cancelada', 'No Realizada') for e in estatuses):
                cell.value = estatuses[0]
                cell.fill = ROJO_LIGHT
            else:
                cell.value = estatuses[0]

            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN

        ws3.cell(row=row, column=6, value=terminadas).font = FONT_NORMAL
        ws3.cell(row=row, column=7, value=total).font = FONT_NORMAL

        cell_clasif = ws3.cell(row=row, column=8, value=clasificacion)
        cell_clasif.font = FONT_BOLD
        if clasificacion == 'COMPLETO':
            cell_clasif.fill = VERDE_LIGHT
        elif clasificacion == 'SIN CITA':
            cell_clasif.fill = AMARILLO_LIGHT
        elif clasificacion in ('NO REALIZADO', 'PARCIAL'):
            cell_clasif.fill = ROJO_LIGHT

        for c in range(1, len(s_headers) + 1):
            ws3.cell(row=row, column=c).border = BORDER_THIN
            ws3.cell(row=row, column=c).alignment = ALIGN_CENTER

        _apply_data_row(ws3, row, len(s_headers), row % 2 == 0)
        row += 1

    _auto_width(ws3)

    # =====================================================================
    # HOJA 4: DETALLE CITAS
    # =====================================================================
    ws4 = wb.create_sheet('Detalle Citas')
    ws4.sheet_properties.tabColor = '0891B2'

    d_headers = ['ID', 'Sucursal', 'Zona', 'Servicio', 'Tipo', 'Mes',
                 'Estatus', 'Técnico', 'Fecha Inicio', 'Fecha Fin']
    for c, h in enumerate(d_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    _apply_header_row(ws4, 1, len(d_headers))

    row = 2
    for r in sorted(limpios, key=lambda x: (x['mes'], x['propiedad'])):
        vals = [
            r['id'], r['propiedad'], r['zona'], r['servicio'],
            r['tipo_servicio'].capitalize(), MESES_Q1.get(r['mes'], ''),
            r['estatus'], r['tecnico'],
            r['fecha_inicio'].strftime('%d/%m/%Y %H:%M') if r['fecha_inicio'] else '',
            r['fecha_terminacion'].strftime('%d/%m/%Y %H:%M') if r['fecha_terminacion'] else '',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=row, column=c, value=v)
            cell.border = BORDER_THIN
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER

        # Colorear estatus
        cell_est = ws4.cell(row=row, column=7)
        if r['estatus'] == 'Terminada':
            cell_est.fill = VERDE_LIGHT
        elif r['estatus'] in ('Cancelada', 'No Realizada'):
            cell_est.fill = ROJO_LIGHT

        _apply_data_row(ws4, row, len(d_headers), row % 2 == 0)
        row += 1

    # Autofiltro
    ws4.auto_filter.ref = f'A1:{get_column_letter(len(d_headers))}{row - 1}'
    _auto_width(ws4)

    # =====================================================================
    # HOJA 5: AUDITORÍA
    # =====================================================================
    ws5 = wb.create_sheet('Auditoría')
    ws5.sheet_properties.tabColor = 'DC2626'

    a_headers = ['ID', 'Sucursal', 'Servicio', 'Mes', 'Estatus',
                 'Fecha Inicio', 'Motivo Descarte']
    for c, h in enumerate(a_headers, 1):
        ws5.cell(row=1, column=c, value=h)
    _apply_header_row(ws5, 1, len(a_headers),
                      PatternFill('solid', fgColor='DC2626'))

    row = 2
    for r in descartados:
        vals = [
            r['id'], r['propiedad'], r['servicio'],
            MESES_Q1.get(r['mes'], ''), r['estatus'],
            r['fecha_inicio'].strftime('%d/%m/%Y %H:%M') if r['fecha_inicio'] else '',
            'Duplicado — se conservó registro más reciente/Terminada',
        ]
        for c, v in enumerate(vals, 1):
            cell = ws5.cell(row=row, column=c, value=v)
            cell.border = BORDER_THIN
            cell.font = FONT_NORMAL
        _apply_data_row(ws5, row, len(a_headers), row % 2 == 0)
        row += 1

    if not descartados:
        ws5.cell(row=2, column=1,
                 value='No se encontraron registros duplicados.')
        ws5.cell(row=2, column=1).font = FONT_SMALL

    _auto_width(ws5)

    # =====================================================================
    # Guardar en BytesIO
    # =====================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
